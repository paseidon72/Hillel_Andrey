import openpyxl
import json
import time

from person import Person


class ColumnException(Exception):
    pass


class DB(object):

    name_fields = ('Name', 'Surname', 'Lastname', 'Birthday', 'Death', 'Sex')
    extensions = {3: ".xlsx", 4: ".xlsx", 5: ".json"}

    def __init__(self):
        self.data = []
        self.choices = {3: self.__load, 4: self.__save, 5: self.__export}

    def __load(self, f_name):
        if f_name:
            try:
                wb = openpyxl.load_workbook(f_name)
                sheet = wb.active
                rows = sheet.max_row
                cols = sheet.max_column
                if cols != 6:
                    raise ColumnException
                self.data = []
                for i in range(2, rows + 1):
                    name = sheet.cell(row=i, column=1).value
                    s_name = sheet.cell(row=i, column=2).value
                    s_name = s_name if s_name is not None else ''
                    l_name = sheet.cell(row=i, column=3).value
                    l_name = l_name if l_name is not None else ''
                    birthday = sheet.cell(row=i, column=4).value
                    death = sheet.cell(row=i, column=5).value
                    death = death if death is not None else ''
                    sex = sheet.cell(row=i, column=6).value
                    person = Person(name, s_name, l_name, birthday, death, sex)
                    if not person.valid():
                        continue
                    self.data.append(person)
                print(f"База данных загружена из в файла {f_name}")
            except FileNotFoundError:
                print(f"Файл {f_name} отсутствует")
            except ColumnException:
                print(f"В файле {f_name} не верное количество столбцов")
            except Exception:
                print(f"При чтении файла {f_name} пошло что-то не так...")
        time.sleep(2)

    def __save(self, f_name):
        if f_name:
            wb = openpyxl.Workbook()
            sheet = wb.active
            for row_index, person in enumerate(self.data):
                values = (person.name, person.s_name, person.l_name,
                          person.birthday, person.death, person.sex)
                for col_index, value in enumerate(values):
                    cell = sheet.cell(row=row_index + 2, column=col_index + 1)
                    if row_index == 0:
                        cell_name = sheet.cell(row=1, column=col_index + 1)
                        cell_name.value = self.name_fields[col_index]
                    cell.value = value
            wb.save(f_name)
            print(f"База данных сохранена в файле {f_name}")
            time.sleep(2)

    def __export(self, f_name):
        if f_name:
            json_data = []
            for person in self.data:
                json_dict = {}
                json_dict[self.name_fields[0]] = person.name
                json_dict[self.name_fields[1]] = person.s_name
                json_dict[self.name_fields[2]] = person.l_name
                json_dict[self.name_fields[3]] = person.birthday
                json_dict[self.name_fields[4]] = person.death
                json_dict[self.name_fields[5]] = person.sex
                json_data.append(json_dict)

            with open(f_name, 'w') as f:
                json.dump(json_data, f)
            print(f"База данных импортирована в файле {f_name}")
            time.sleep(2)

    def input_data(self, *args):
        while True:
            name = input("Введите имя: ")
            s_name = input("Введите фамилию или ENTER: ")
            l_name = input("Введите отчество или ENTER: ")
            birthday = input("Введите дату рождения (20.10.1980): ")
            death = input("Введите дату смерти (20.10.1980) или ENTER: ")
            sex = input("Введите пол (m/f): ")
            person = Person(name, s_name, l_name, birthday, death, sex)
            if not person.valid():
                print("Ошибка ввода:", person.error)
                answer = input("Хотите повторить ввод (д/y): ")
                if answer.lower() not in ("y", "д"):
                    break
                continue
            self.data.append(person)
            print('-' * 50)
            answer = input("Хотите продолжить ввод (д/y): ")
            if answer.lower() not in ("y", "д"):
                break

    def get_find(self, find_word):
        result = []
        for person in self.data:
            if find_word.lower() in person.full_name.lower():
                result.append(person)
        return result

    def find(self, *args):
        while True:
            find_word = input("Введите имя и/или фамилию и/или отчество: ")
            result = self.get_find(find_word)
            if result:
                for person in result:
                    print(person)
            else:
                print("По вашему запросу ничего не найдено")
            print('-' * 50)
            answer = input("Хотите повторить поиск (д/y): ")
            if answer.lower() not in ("y", "д"):
                break

    def file(self, choice):
        f_name = input("Введите имя файла: ")
        f_name = f"{f_name.split('.')[0]}{self.extensions[choice]}"
        self.choices[choice](f_name)