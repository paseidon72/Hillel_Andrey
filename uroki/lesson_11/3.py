# Используя данные из прошлого примера,
# сначала записать их в файл Excel-формата,
# а затем считать из сохранённого файла


import openpyxl


name_of_fields = ["Name", "Class", "Age"]
field_01 = ["Jan", "3", "10"]
field_02 = ["Alex", "5", "12"]
field_03 = ["Michael", "11", "18"]


wb = openpyxl.Workbook()
print(wb.sheetnames)

wb.create_sheet(title='Первый лист', index=0)
print(wb.sheetnames)

wb.remove(wb['Sheet'])
print(wb.sheetnames)

sheet = wb['Первый лист']
print(sheet)

for row_index, row in enumerate((name_of_fields, field_01, field_02, field_03)):
    for col_index, value in enumerate(row):
        cell = sheet.cell(row=row_index+1, column=col_index+1)
        cell.value = value

wb.save('task_04.xlsx')

wb2 = openpyxl.load_workbook('task_05.xlsx')
print(wb2.sheetnames)

sheet = wb2.active
print(sheet)

print(sheet['A2'].value)

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

data = []

for i in range(1, rows + 1):
    row_data = []
    for j in range(1, cols + 1):
        cell = sheet.cell(row=i, column=j)
        row_data.append(cell.value)
    data.append(row_data)

print(data)