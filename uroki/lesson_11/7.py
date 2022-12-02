# Прочитать сохранённый csv-файл из задания №18 и сохранить данные
# в excel-файл, кроме возраста – столбец с этими данными не нужен.


import csv
import openpyxl

input_data = []
names = None

# открытие файла на чтение
with open("file_csv.csv", encoding='utf-8') as r_file:
    file_reader = csv.reader(r_file, delimiter=",")
    count = 0
    for row in file_reader:
        if count == 0:
            names = row  # вычитывание наименований столбцов (первая строка в файле)
        else:
            input_data.append(row)  # вычитывание данных построчно
        count += 1

# нахождение индекса столбца 'age'
age_index = [index for index, item in enumerate(names) if item == 'age']
age_index = age_index[0] if age_index else None

# если этот столбец есть, то удалить его значение и со списка наименований и со списка с данными
if age_index is not None:
    names.pop(age_index)
    for item_list in input_data:
        item_list.pop(age_index)

wb = openpyxl.Workbook()  # открытие книги excel
sheet = wb.active  # выбор активного листа

# запись верхней строки person 1, person 2 ... начиная со второго столбца
# кол-во записей person равняется количеству списков внутри input_data
for item in range(1, len(input_data) + 1):
    cell = sheet.cell(row=1, column=item + 1)
    cell.value = f'person {item}'

# цикл перебора строк
for name_index, name in enumerate(names):
    cell = sheet.cell(row=name_index+2, column=1)
    cell.value = name
    # цикл перебора столбцов
    for item in range(1, len(input_data) + 1):
        cell = sheet.cell(row=name_index+2, column=item + 1)
        cell.value = input_data[item - 1][name_index]

wb.save('file_xlsx.xlsx')