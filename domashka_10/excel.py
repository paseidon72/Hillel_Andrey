import openpyxl


name_file = ['id', 'firstname', 'age', 'phone']
file_1 = ['123456', 'name', 22, +380501234567]
file_2 = ['789654', 'name1', 25, +380507654321]
file_3 = ['321465', 'name2', 30, +380509876543]
file_4 = ['135802', 'name3', 32, +380503456789]
file_5 = ['133456', 'name4', 38, +380501357937]

wb = openpyxl.Workbook()


wb.create_sheet(title='Первый лист', index=0)
wb.remove('sheet')
sheet = wb['Первый лист']

for row_index, row in enumerate((name_file, file_1, file_2, file_3, file_4, file_5)):
    for col_index, value in enumerate(row):
        cell = sheet.cell(row=row_index + 1, column=col_index + 1)
        cell.value = value

wb.save('task_3.xlsx')
